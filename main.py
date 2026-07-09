# Dependencies:
# pip install groq python-dotenv rich openpyxl

import csv
import json
import os
import smtplib
import sys
import time
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from dotenv import load_dotenv
from groq import Groq
from openpyxl import load_workbook
from ddgs import DDGS
from rich.align import Align
from rich.box import DOUBLE, HEAVY, ROUNDED
from rich.columns import Columns
from rich.console import Console, Group
from rich.layout import Layout
from rich.live import Live
from rich.panel import Panel
from rich.progress import (
    BarColumn,
    Progress,
    SpinnerColumn,
    TaskProgressColumn,
    TextColumn,
    TimeElapsedColumn,
    TimeRemainingColumn,
)
from rich.table import Table
from rich.text import Text

from src.outreach.prompts import HR_SYSTEM_PROMPT, EMAIL_CRITIC_PROMPT
from src.config.config import Config

BATCH_SIZE = 100


# ── Configuration ──────────────────────────────────────────────────────────────

def load_config() -> dict:
    """Load and validate environment variables from .env file."""
    load_dotenv()
    required = ["GMAIL_ADDRESS", "GMAIL_APP_PASSWORD", "GROQ_API_KEY", "RESUME_PATH"]
    config = {}
    missing = []
    for key in required:
        val = os.getenv(key)
        if not val:
            missing.append(key)
        else:
            config[key] = val
    if missing:
        console = Console()
        console.print(
            f"[bold red]Missing environment variables: {', '.join(missing)}[/bold red]\n"
            "Create a .env file using .env.example as a template."
        )
        sys.exit(1)

    script_dir = Path(__file__).resolve().parent
    resume = Path(config["RESUME_PATH"])
    if not resume.is_absolute():
        resume = (script_dir / resume).resolve()

    if not resume.exists():
        fallback = None
        lower_name = resume.name.lower()
        candidates = [
            p
            for p in script_dir.glob("*.pdf")
            if lower_name in p.name.lower() or "resume" in p.name.lower()
        ]
        if len(candidates) == 1:
            fallback = candidates[0]
        elif len(candidates) > 1:
            fallback = next((p for p in candidates if p.name.lower() == lower_name), None)

        if fallback:
            resume = fallback

    if not resume.exists():
        Console().print(f"[bold red]Resume file not found: {resume}[/bold red]")
        sys.exit(1)

    config["RESUME_PATH"] = str(resume)
    return config


# ── Read Leads ─────────────────────────────────────────────────────────────────

def read_leads(filepath: str = "data/leads.xlsx", max_rows: int = BATCH_SIZE) -> list[dict]:
    """Read up to max_rows leads from an Excel file."""
    path = Path(filepath)
    if not path.exists():
        Console().print(f"[bold red]Leads file not found: {path}[/bold red]")
        sys.exit(1)

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        Console().print("[bold red]Leads file is empty.[/bold red]")
        sys.exit(1)

    header = [str(cell).strip().lower() if cell else "" for cell in rows[0]]

    # Support multiple common header names
    company_aliases = ["company_name", "company"]
    email_aliases = ["hr_email", "email"]
    col_company = next((header.index(a) for a in company_aliases if a in header), None)
    col_email = next((header.index(a) for a in email_aliases if a in header), None)

    if col_company is None or col_email is None:
        Console().print(
            "[bold red]Leads file must have a company column (company_name/company) "
            "and an email column (hr_email/email).[/bold red]"
        )
        sys.exit(1)

    leads = []
    for row in rows[1:]:
        if len(leads) >= max_rows:
            break
        company = str(row[col_company]).strip() if row[col_company] else ""
        email = str(row[col_email]).strip() if row[col_email] else ""
        if company and email:
            leads.append({"company_name": company, "hr_email": email})

    wb.close()
    return leads


def delete_processed_rows(filepath: str, count: int) -> int:
    """Delete the first `count` data rows from the Excel file. Returns remaining row count."""
    wb = load_workbook(filepath)
    ws = wb.active
    for _ in range(count):
        if ws.max_row < 2:
            break
        ws.delete_rows(2)
    remaining = max(ws.max_row - 1, 0)  # subtract header
    wb.save(filepath)
    wb.close()
    return remaining


# ── Groq Email Generation ─────────────────────────────────────────────────────

def get_company_context(company_name: str) -> str:
    """Perform a web search to get context on what the company does."""
    try:
        from ddgs import DDGS
        query = f"{company_name} company what they do"
        results = DDGS().text(query, max_results=1)
        if results and len(results) > 0:
            return results[0].get("body", "No context found.")
    except Exception as e:
        pass
    return "A company operating in the technology space."

def generate_email(client: Groq, company_name: str, context: str) -> tuple[str, str]:
    """Use Groq LLM to generate a personalized email subject and body with Critic Loop."""
    try:
        with open(Config.DATA_DIR / "context" / "yash_master_profile.md", "r") as f:
            yash_master_profile = f.read()
        context_str = f"\n\n--- YASH MASTER PROFILE ---\n{yash_master_profile}"
    except Exception:
        context_str = ""

    strategy = {
        "selected_project": "GDSC Search Engine",
        "rationale": "GDSC Search Engine demonstrates building semantic retrieval and LLM backends.",
        "company_observation": context
    }
    
    subject_variant = f"IIT Roorkee Student | AI / Software Engineering"
    
    max_attempts = 3
    email_data = {}
    critic_data = {}
    
    for attempt in range(max_attempts):
        try:
            response = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[
                    {"role": "system", "content": HR_SYSTEM_PROMPT + context_str},
                    {"role": "user", "content": f"Target Company: {company_name}\nJob Role: AI / Software Engineer\nStrategy: {json.dumps(strategy)}\nForced Subject: {subject_variant}"}
                ],
                temperature=0.3,
                response_format={"type": "json_object"}
            )
            email_data = json.loads(response.choices[0].message.content)
            draft_body = email_data.get("body", "")
            
            critic_response = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[
                    {"role": "system", "content": EMAIL_CRITIC_PROMPT},
                    {"role": "user", "content": f"Please evaluate the following email draft:\n\n{draft_body}"}
                ],
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            critic_data = json.loads(critic_response.choices[0].message.content)
            
            if critic_data.get("status") == "PASS":
                break
        except Exception as e:
            if "429" in str(e) and getattr(Config, "GROQ_API_KEY_BACKUP", ""):
                client = Groq(api_key=Config.GROQ_API_KEY_BACKUP)
                continue
            raise e
            
    if critic_data.get("status") != "PASS":
        raise Exception(f"Email Critic failed after {max_attempts} attempts. Feedback: {critic_data.get('feedback', 'No feedback (e.g. rate limit)')}")
        
    return email_data.get("subject", subject_variant), email_data.get("body", "")


# ── Email Sending ──────────────────────────────────────────────────────────────

def send_email(
    sender: str,
    password: str,
    to_email: str,
    subject: str,
    body: str,
    resume_path: str,
) -> None:
    """Send an email with a PDF resume attachment via Gmail SMTP."""
    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = to_email
    msg["Subject"] = subject
    # Normalize any double newlines down to single, then convert to HTML paragraphs
    body = body.replace("\r\n", "\n").strip()
    while "\n\n" in body:
        body = body.replace("\n\n", "\n")
    paragraphs = body.split("\n")
    html_body = "".join(f"<p style='margin:0 0 10px 0;'>{p}</p>" for p in paragraphs if p.strip())
    html_content = f"""\
<html>
<body style="font-family: Arial, sans-serif; font-size: 14px; line-height: 1.5; color: #222;">
{html_body}
</body>
</html>"""
    msg.attach(MIMEText(html_content, "html"))

    resume = Path(resume_path)
    with open(resume, "rb") as f:
        attachment = MIMEApplication(f.read(), _subtype="pdf")
        attachment.add_header(
            "Content-Disposition", "attachment", filename=resume.name
        )
        msg.attach(attachment)

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(sender, password)
        server.sendmail(sender, to_email, msg.as_string())


# ── Logging ────────────────────────────────────────────────────────────────────

def init_log(path: str = "logs.csv") -> None:
    """Create the log CSV with headers."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["company_name", "hr_email", "status", "error_message", "timestamp"])


def append_log(
    company: str, email: str, status: str, error: str, ts: str, path: str = "logs.csv"
) -> None:
    """Append a single log row."""
    with open(path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([company, email, status, error, ts])


# ── Dashboard ──────────────────────────────────────────────────────────────────

BANNER = r"""
  __    __  ____     ____  _   _  _____  ____    ___    _    ___  _   _ 
 /__\  /__\/  _ \   /  _ \| | | ||_   _||  _ \  / _ \  / \  / __|| | | |
/_\\\\//_\\\\| |_) ) | | | || | | |  | |  | |_) )| |_| |/ _ \| |   | |_| |
//__ //__ \\  _ <  | |_| || |_| |  | |  |  _ <\____  / ___ \ |__ |  _  |
\__/ \__/|_| \_\  \___/  \___/   |_|  |_| \_\    |_/_/   \_\___||_| |_|
"""


def render_banner() -> Panel:
    """Vibrant gradient ASCII banner."""
    palette = [
        "bold magenta", "bold red", "bold yellow", "bold green",
        "bold cyan", "bold blue", "bold bright_magenta",
    ]
    text = Text()
    lines = BANNER.strip("\n").splitlines()
    for i, line in enumerate(lines):
        text.append(line + "\n", style=palette[i % len(palette)])
    subtitle = Text(
        "✨ HR Cold Outreach — Automated Email Dispatcher ✨",
        style="bold white on magenta",
        justify="center",
    )
    return Panel(
        Align.center(Group(Align.center(text), Text(""), subtitle)),
        box=DOUBLE,
        border_style="bright_magenta",
        padding=(0, 2),
    )


def make_table() -> Table:
    """Create the colorful status table."""
    table = Table(
        title="📨 [bold magenta]Email Dispatch Log[/bold magenta]",
        expand=True,
        box=ROUNDED,
        border_style="bright_blue",
        header_style="bold white on blue",
        row_styles=["", "on grey7"],
        show_lines=False,
    )
    table.add_column("#", style="bold cyan", width=5, justify="right")
    table.add_column("Company", style="bold yellow")
    table.add_column("Email", style="bright_white")
    table.add_column("Status", justify="center", width=14)
    table.add_column("Time", style="dim italic", width=10)
    return table


def make_stats_panel(sent: int, failed: int, pending: int, total: int, current: str) -> Panel:
    """Live stats grid with colorful counters."""
    grid = Table.grid(expand=True, padding=(0, 1))
    grid.add_column(justify="center", ratio=1)
    grid.add_column(justify="center", ratio=1)
    grid.add_column(justify="center", ratio=1)
    grid.add_column(justify="center", ratio=1)

    def cell(label, value, color):
        return Panel(
            Align.center(
                Group(
                    Text(str(value), style=f"bold {color}", justify="center"),
                    Text(label, style="dim white", justify="center"),
                )
            ),
            border_style=color,
            box=HEAVY,
            padding=(0, 1),
        )

    grid.add_row(
        cell("TOTAL", total, "bright_cyan"),
        cell("SENT ✅", sent, "bright_green"),
        cell("FAILED ❌", failed, "bright_red"),
        cell("PENDING ⏳", pending, "bright_yellow"),
    )

    current_text = Text()
    current_text.append("▶ Now contacting: ", style="bold bright_magenta")
    current_text.append(current or "—", style="bold bright_white on purple")

    return Panel(
        Group(grid, Text(""), Align.center(current_text)),
        title="[bold bright_white on blue] LIVE STATUS [/bold bright_white on blue]",
        border_style="bright_magenta",
        box=DOUBLE,
    )


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    console = Console()
    console.print(render_banner())

    # 1. Load config & leads (max BATCH_SIZE per run)
    with console.status("[bold cyan]Loading configuration & leads…", spinner="dots12"):
        config = load_config()
        leads = read_leads()

    if not leads:
        console.print(Panel(
            "[bold yellow]📦 No leads remaining in leads.xlsx. Nothing to send.[/bold yellow]",
            border_style="yellow", box=DOUBLE,
        ))
        return

    total = len(leads)
    console.print(Panel(
        f"[bold bright_green]✅ Loaded [bright_white]{total}[/bright_white] lead(s) from "
        f"leads.xlsx[/bold bright_green]  [dim](batch up to {BATCH_SIZE})[/dim]",
        border_style="green", box=ROUNDED, expand=False,
    ))

    # 2. Init Groq client & log file
    groq_client = Groq(api_key=config["GROQ_API_KEY"])
    init_log()

    # 3. Prepare colorful dashboard components
    progress = Progress(
        SpinnerColumn(spinner_name="dots12", style="bold magenta"),
        TextColumn("[bold bright_cyan]🚀 Dispatching"),
        BarColumn(
            bar_width=None,
            complete_style="bright_green",
            finished_style="bold green",
            pulse_style="bright_magenta",
        ),
        TaskProgressColumn(),
        TextColumn("[bold yellow]{task.completed}[/bold yellow]/[bold cyan]{task.total}[/bold cyan]"),
        TextColumn("• [dim]elapsed[/dim]"),
        TimeElapsedColumn(),
        TextColumn("• [dim]eta[/dim]"),
        TimeRemainingColumn(),
        expand=True,
    )
    task_id = progress.add_task("emails", total=total)
    table = make_table()

    sent_count = 0
    fail_count = 0
    start_time = time.time()
    current_company = "starting…"

    def render_dashboard() -> Group:
        pending = total - sent_count - fail_count
        return Group(
            make_stats_panel(sent_count, fail_count, pending, total, current_company),
            Panel(progress, border_style="bright_cyan", box=ROUNDED,
                  title="[bold bright_white]📡 Progress[/bold bright_white]"),
            table,
        )

    # 4. Live dashboard loop
    with Live(render_dashboard(), console=console, refresh_per_second=10, screen=False) as live:
        for idx, lead in enumerate(leads, start=1):
            company = lead["company_name"]
            hr_email = lead["hr_email"]
            ts = datetime.now().strftime("%H:%M:%S")
            error_msg = ""
            current_company = company
            live.update(render_dashboard())

            try:
                current_company = f"Searching: {company}"
                live.update(render_dashboard())
                context = get_company_context(company)
                
                current_company = f"Generating: {company}"
                live.update(render_dashboard())
                subject, body = generate_email(groq_client, company, context)
                
                current_company = f"Sending: {company}"
                live.update(render_dashboard())
                send_email(
                    sender=config["GMAIL_ADDRESS"],
                    password=config["GMAIL_APP_PASSWORD"],
                    to_email=hr_email,
                    subject=subject,
                    body=body,
                    resume_path=config["RESUME_PATH"],
                )
                status_label = "[bold black on bright_green] ✅ SENT [/bold black on bright_green]"
                status_csv = "sent"
                sent_count += 1
            except Exception as exc:
                status_label = "[bold white on red] ❌ FAILED [/bold white on red]"
                status_csv = "failed"
                error_msg = str(exc)
                fail_count += 1

            # Update dashboard
            table.add_row(str(idx), company, hr_email, status_label, ts)
            progress.update(task_id, advance=1)
            live.update(render_dashboard())

            # Persist log
            append_log(company, hr_email, status_csv, error_msg, ts)

            # Rate-limit pause (skip after last email)
            if idx < total:
                time.sleep(3)

        current_company = "🏁 complete"
        live.update(render_dashboard())

    # 5. Summary
    elapsed = time.time() - start_time
    minutes, seconds = divmod(int(elapsed), 60)
    success_rate = (sent_count / total * 100) if total else 0

    summary_table = Table.grid(padding=(0, 2))
    summary_table.add_column(style="bold bright_white", justify="right")
    summary_table.add_column()
    summary_table.add_row("📤 Total attempted:", f"[bold bright_cyan]{total}[/bold bright_cyan]")
    summary_table.add_row("✅ Sent:", f"[bold bright_green]{sent_count}[/bold bright_green]")
    summary_table.add_row("❌ Failed:", f"[bold bright_red]{fail_count}[/bold bright_red]")
    summary_table.add_row("📊 Success rate:", f"[bold yellow]{success_rate:.1f}%[/bold yellow]")
    summary_table.add_row("⏱️  Time taken:", f"[bold bright_magenta]{minutes}m {seconds}s[/bold bright_magenta]")

    console.print()
    console.print(Panel(
        Align.center(summary_table),
        title="[bold bright_white on magenta] 🎉 DISPATCH SUMMARY 🎉 [/bold bright_white on magenta]",
        border_style="bright_magenta",
        box=DOUBLE,
        padding=(1, 4),
        expand=False,
    ))
    console.print(f"\n[dim italic]📄 Log saved to logs.csv[/dim italic]")

    # 6. Delete processed rows from leads.xlsx
    remaining = delete_processed_rows("data/leads.xlsx", total)
    console.print(Panel(
        f"[bold]🗑️  Deleted [bright_red]{total}[/bright_red] processed row(s) — "
        f"[bright_yellow]{remaining}[/bright_yellow] lead(s) remaining in leads.xlsx[/bold]",
        border_style="bright_blue", box=ROUNDED, expand=False,
    ))
    if remaining == 0:
        console.print(Panel(
            "[bold bright_green]🎊 ALL LEADS HAVE BEEN PROCESSED! No more emails to send. 🎊[/bold bright_green]",
            border_style="bright_green", box=DOUBLE, expand=False,
        ))


if __name__ == "__main__":
    main()
